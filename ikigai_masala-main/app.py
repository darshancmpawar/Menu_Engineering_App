"""
Streamlit frontend for Ikigai Masala Menu Planning.

Single entry point - auto-starts the Flask API backend in a background thread.

Run with:
    cd ikigai_masala-main
    streamlit run app.py
"""

import os
import sys

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)
os.chdir(_APP_DIR)

import datetime as dt
import html
import io
import logging
import re
import threading
import time

import streamlit as st

from ui.api_client import MenuApiClient, RuleDiagnosticsBlockedError
from ui.formatters import (
    display_label_for_slot_id,
    flatten_api_solution,
    format_item_for_ui,
    format_item_html,
    nonveg_slots_from_solution,
    slot_sort_key,
    THEME_TAG_COLORS,
    THEME_ICONS,
)
from ui.styles import STYLES
from ui.branding import favicon as _favicon, logo_img_tag
from ui.backend_probe import health_check, pick_backend_port
from customisation.main import render_customisation_editor


logger = logging.getLogger(__name__)

# Streamlit-side cap on solver wall-clock per request. The API itself
# accepts up to MAX_TIME_LIMIT_SECONDS; we send a tighter value so a
# single slow request doesn't pin a worker for the full 10-minute API
# ceiling. Tuned against the 5-day default plan; revisit if num_days
# grows or rule count balloons.
_PLANNING_TIME_LIMIT_SECONDS = 180


def _render_view_error(view_name: str, exc: BaseException) -> None:
    """Show a clean Streamlit-native error block for an unhandled
    exception inside a top-level view (editor / planner).

    The full traceback lands in the server log via ``logger.exception``;
    the user sees a short message + a button to bounce back to the
    planner. Without this guard a render-side bug renders a half-page
    or — depending on Streamlit's config — a full Python traceback,
    neither of which is acceptable for a multi-user deployment.
    """
    logger.exception("Unhandled error in %s view", view_name)
    st.error(
        f"Something went wrong loading the {view_name}. "
        "The error has been logged. Please go back and try again."
    )
    if st.button("Back to planner", key=f"err_back_{view_name}"):
        st.session_state.view = "planner"
        st.rerun()


# ---------------------------------------------------------------------------
# Auto-start Flask API backend
# ---------------------------------------------------------------------------
_BACKEND_URL = None  # set by _ensure_backend_running()


def _start_flask_backend(port: int) -> None:
    # api.app's module-level validate_required_env() raises if any
    # required var is missing; let that bubble up so the Streamlit
    # process shows a clear error instead of a silent backend crash.
    # Logging is configured inside api.app via configure_logging(),
    # so don't install a second root handler here.
    from api.app import app as flask_app
    flask_app.run(host="127.0.0.1", port=port, debug=False,
                  use_reloader=False, threaded=True)


def _ensure_backend_running() -> str:
    """Start the backend if needed and return its base URL.

    Raises ``RuntimeError`` if no port is available or the backend does
    not become healthy within the startup window — the caller should
    surface the error to the user rather than hit an unrelated service
    that happens to sit on port 5000.
    """
    global _BACKEND_URL
    port = pick_backend_port()
    url = f"http://localhost:{port}"
    if health_check(port):
        _BACKEND_URL = url
        return url
    if "flask_started" not in st.session_state:
        t = threading.Thread(
            target=_start_flask_backend, args=(port,), daemon=True,
        )
        t.start()
        st.session_state.flask_started = True
    for _ in range(20):
        if health_check(port):
            _BACKEND_URL = url
            return url
        time.sleep(0.5)
    raise RuntimeError(
        f"Backend did not become healthy on port {port} within 10s. "
        "Check the Streamlit server logs for errors."
    )


# ---------------------------------------------------------------------------
# Page config — MUST be first Streamlit command
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Ikigai Masala - Menu Planner",
    page_icon=_favicon(),  # SmartQ logo if ui/assets/smartq_logo.png exists, else emoji
    layout="wide",
    initial_sidebar_state="expanded",
)

# The planner uses the dark theme (ui/styles.py); the customisation editor is
# a self-contained full-page view rendered against the Pulse light theme, so
# skip the dark stylesheet while the editor is active — otherwise the two sets
# of !important rules fight and the editor renders half-dark. The editor
# injects its own Pulse CSS in render_customisation_editor().
if st.session_state.get("view", "planner") != "editor":
    st.markdown(STYLES, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Backend must be up before we render anything (the UI hits the API).
# ---------------------------------------------------------------------------
# Validate required env vars first so a misconfigured deployment shows a
# clear Streamlit-native error instead of "backend did not become healthy"
# after the spawned thread silently crashes.
try:
    from api.config import validate_required_env
    validate_required_env()
except RuntimeError as e:
    st.error(str(e))
    st.stop()

try:
    _ensure_backend_running()
except RuntimeError as e:
    st.error(str(e))
    st.stop()

# Streamlit reruns the entire script on every widget interaction, so a
# naïve `MenuApiClient(_BACKEND_URL)` rebuilds a fresh requests.Session
# on every click — connection pool, retry adapter, etc. — for no
# reason. Cache one client per base URL so the underlying HTTPS pool
# survives across reruns.
@st.cache_resource(show_spinner=False)
def _get_api_client(base_url: str) -> MenuApiClient:
    return MenuApiClient(base_url)


client = _get_api_client(_BACKEND_URL)


# Cache low-churn reads (60s TTL): the sidebar's client picker re-renders
# on every interaction but the list itself only changes when someone
# creates/deletes a client. The underscore-prefixed arg is excluded from
# the hash key (Streamlit can't hash MenuApiClient).
@st.cache_data(ttl=60, show_spinner=False)
def _cached_list_clients(_api: MenuApiClient) -> list:
    """Return ``[{'name', 'city'}, …]`` for the sidebar's client + city pickers."""
    return _api.list_clients_with_city()

# ---------------------------------------------------------------------------
# Session state initialization
# ---------------------------------------------------------------------------
_SESSION_DEFAULTS = {
    "plan": None,
    "plan_dates": [],
    "day_types": {},
    "pool_warnings": [],
    "client_name": None,
    "changes_log": [],
    "view": "planner",
    # "history" when the current plan was loaded from /saved-plan,
    # "solver" when it came from /plan, "modified" once the user has
    # regenerated a cell (so the on-screen plan no longer matches the
    # DB version), "preflight_blocked" when the diagnostic gate stopped
    # the solver from running. Drives the badge on the page header.
    "plan_source": None,
    # Pre-flight rule_diagnostics from the most recent /plan or
    # /saved-plan response (or from a RuleDiagnosticsBlockedError).
    # Empty list = nothing to show. Rendered as the inline expander
    # above the plan table.
    "rule_diagnostics": [],
    "diagnostics_summary": None,
    # Unified plan state: a list of per-counter "blocks". Single-cuisine
    # clients have one block; multi-cuisine clients have one per counter.
    # Each block: {name, plan, plan_dates, day_types, pool_warnings, source, error}.
    "plan_blocks": [],
    "plan_mode": "single",
}
for key, default in _SESSION_DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ---------------------------------------------------------------------------
# Editor view (full-page)
# ---------------------------------------------------------------------------
if st.session_state.view == "editor":
    try:
        render_customisation_editor(client)
    except Exception as _exc:
        # A Supabase blip while loading client config, a malformed rule
        # in client_rules.json, etc. Log + show a friendly fallback so
        # the user can navigate out instead of staring at a half-page.
        _render_view_error("editor", _exc)
    st.stop()

# ---------------------------------------------------------------------------
# Sidebar (planner view)
# ---------------------------------------------------------------------------
with st.sidebar:
    _logo_tag = logo_img_tag(height=30)
    _brand_icon = (
        f'<div class="sidebar-brand-icon" style="background:transparent;'
        f'box-shadow:none;">{_logo_tag}</div>'
        if _logo_tag
        else '<div class="sidebar-brand-icon">&#127835;</div>'
    )
    st.markdown(f"""<div class="sidebar-brand">
        <div class="sidebar-brand-row">
            {_brand_icon}
            <div>
                <h2>Ikigai Masala</h2>
                <p>Weekly Menu Planner</p>
            </div>
        </div>
    </div>""", unsafe_allow_html=True)

    try:
        clients_detail = _cached_list_clients(client)
    except (ConnectionError, OSError, ValueError):
        clients_detail = []
        st.error("Cannot reach API.")

    # City filter — single-select, default "All". Only cities that actually
    # have clients are offered, so no selection ever yields an empty list.
    cities = sorted({c.get("city") for c in clients_detail if c.get("city")})
    city_filter = st.selectbox("City", ["All"] + cities,
                               key="planner_city_filter")
    if city_filter == "All":
        clients_list = [c["name"] for c in clients_detail]
    else:
        clients_list = [c["name"] for c in clients_detail
                        if c.get("city") == city_filter]

    selected_client = st.selectbox("Client",
        clients_list if clients_list else ["(no clients)"],
        key="planner_client_select")
    start_date = st.date_input("Start date", value=dt.date.today(),
                               key="planner_start_date")
    num_days = st.slider("Weekdays", min_value=1, max_value=20, value=5,
                         key="planner_num_days",
                         help="Number of weekdays (Sat/Sun are skipped)")

    st.divider()
    generate_clicked = st.button("Generate Menu Plan", type="primary",
                                 key="planner_generate_btn",
                                 use_container_width=True)

# ---------------------------------------------------------------------------
# Page header
# ---------------------------------------------------------------------------
_hdr_col1, _hdr_col2 = st.columns([5, 2])
_SOURCE_BADGES = {
    # bg, fg, label, title attr — Pulse low-emphasis badge tints
    "history":  ("#E5FFF1", "#1AA45B", "Loaded from history",
                 "These exact dates already had a saved plan — shown as-is."),
    "solver":   ("#F3ECFF", "#6F42C1", "Freshly generated",
                 "No saved plan for these dates — solver produced this from scratch."),
    "modified": ("#FFF5E8", "#C56A00", "Modified — unsaved",
                 "You regenerated at least one cell since this plan was loaded."),
    "preflight_blocked": ("#FFE7E9", "#C40D1B", "Pre-flight blocked",
                 "Diagnostic checks found a guaranteed failure; solver skipped."),
}


# bg, fg, label per Diagnostic severity (Pulse status tints).
_SEVERITY_STYLE = {
    "error":   ("#FFE7E9", "#C40D1B", "Error"),
    "warning": ("#FFF5E8", "#C56A00", "Warning"),
    "info":    ("#EBF3FF", "#0D6EFD", "Info"),
}


def _render_diagnostics_expander(diagnostics, summary):
    """Render the inline 'Diagnostics' expander above the plan table.

    Auto-expanded when any error is present (the user must act);
    collapsed otherwise. Sectioned by severity so errors are visible
    first. Reuses the design tokens from ``ui/styles.py``.
    """
    if not diagnostics:
        return
    has_error = bool(summary and summary.get("errors", 0)) or any(
        d.get("severity") == "error" for d in diagnostics
    )
    counts = []
    if summary:
        if summary.get("errors"):
            counts.append(f"{summary['errors']} error"
                          f"{'s' if summary['errors'] != 1 else ''}")
        if summary.get("warnings"):
            counts.append(f"{summary['warnings']} warning"
                          f"{'s' if summary['warnings'] != 1 else ''}")
        if summary.get("infos"):
            counts.append(f"{summary['infos']} info")
    label = (
        f"Diagnostics ({', '.join(counts)})" if counts else "Diagnostics"
    )

    with st.expander(label, expanded=has_error):
        # Group by severity so errors come first regardless of how the
        # server sorted them.
        order = ("error", "warning", "info")
        grouped = {sev: [d for d in diagnostics if d.get("severity") == sev] for sev in order}
        for sev in order:
            items = grouped[sev]
            if not items:
                continue
            bg, fg, sev_label = _SEVERITY_STYLE.get(sev, ("#F0F0F0", "#555555", sev.title()))
            st.markdown(
                f'<p style="font-size:0.85rem;font-weight:700;color:{fg};'
                f'margin:0.5rem 0 0.4rem;">{sev_label}'
                f' ({len(items)})</p>',
                unsafe_allow_html=True,
            )
            for d in items:
                rule_pill = html.escape(d.get("rule_type") or d.get("rule") or "?")
                msg = html.escape(d.get("message") or "")
                suggestion = html.escape(d.get("suggestion") or "")
                affected = d.get("affected") or {}
                chips = []
                # Surface the most commonly-useful affected fields as
                # chips; everything else stays inside ``affected`` for
                # the API surface but isn't visualised.
                for k in ("date", "day_type", "slot"):
                    if k in affected:
                        chips.append(
                            f'<span style="background:#F0F0F0;color:#555555;'
                            f'border-radius:99px;padding:1px 8px;font-size:0.65rem;'
                            f'margin-right:4px;">{html.escape(str(affected[k]))}</span>'
                        )
                chip_html = ''.join(chips)
                st.markdown(
                    f'<div style="background:{bg};border-left:3px solid {fg};'
                    f'padding:0.55rem 0.8rem;border-radius:8px;'
                    f'margin-bottom:0.45rem;">'
                    f'<div style="display:flex;align-items:center;gap:0.45rem;'
                    f'margin-bottom:0.2rem;">'
                    f'<span style="background:{fg};color:{bg};font-weight:700;'
                    f'font-size:0.6rem;letter-spacing:0.04em;text-transform:uppercase;'
                    f'padding:1px 7px;border-radius:99px;">{rule_pill}</span>'
                    f'{chip_html}'
                    f'</div>'
                    f'<div style="color:#131313;font-size:0.85rem;'
                    f'line-height:1.4;">{msg}</div>'
                    + (f'<div style="color:#777777;font-size:0.75rem;'
                       f'margin-top:0.25rem;">Fix: {suggestion}</div>'
                       if suggestion else '')
                    + '</div>',
                    unsafe_allow_html=True,
                )


with _hdr_col1:
    st.markdown('<p class="page-title">Menu Plan</p>', unsafe_allow_html=True)
    if st.session_state.client_name:
        src = st.session_state.get("plan_source")
        badge_html = ""
        if src in _SOURCE_BADGES:
            bg, fg, label, title_attr = _SOURCE_BADGES[src]
            badge_html = (
                f'<span class="plan-source-badge" '
                f'title="{html.escape(title_attr)}" '
                f'style="display:inline-block;margin-left:0.6rem;'
                f'padding:2px 10px;border-radius:99px;font-size:0.7rem;'
                f'font-weight:700;letter-spacing:0.04em;text-transform:uppercase;'
                f'background:{bg};color:{fg};vertical-align:middle;">'
                f'{html.escape(label)}</span>'
            )
        st.markdown(
            f'<p class="page-subtitle">Generated plan for '
            f'{html.escape(st.session_state.client_name)}{badge_html}</p>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            '<p class="page-subtitle">Select a client and generate a plan to get started</p>',
            unsafe_allow_html=True)
with _hdr_col2:
    if st.button("Edit Logic", key="open_editor_btn", use_container_width=True):
        st.session_state.view = "editor"
        st.rerun()

# ---------------------------------------------------------------------------
# Planner render helpers (shared by the single- and multi-counter paths)
# ---------------------------------------------------------------------------
def _flatten_result(result: dict) -> dict:
    """Turn a /plan or /saved-plan response into a plan "block"."""
    solution = result.get("solution", {})
    flat, day_types = flatten_api_solution(solution)
    return {
        "plan": flat,
        "plan_dates": sorted(flat.keys()),
        "day_types": day_types,
        "nonveg": nonveg_slots_from_solution(solution),
        "pool_warnings": result.get("pool_warnings", []),
        "source": "solver",
        "error": None,
    }


def _date_label(d_str: str) -> str:
    try:
        return dt.date.fromisoformat(d_str).strftime("%a %d %b")
    except ValueError:
        return d_str


def _menu_table_html(plan: dict, plan_dates: list, day_types: dict,
                     nonveg: dict | None = None) -> str:
    nonveg = nonveg or {}
    header_html = '<tr><th>Category</th>'
    for d_str in plan_dates:
        try:
            d_lbl = dt.date.fromisoformat(d_str).strftime("%a %d %b")
        except ValueError:
            d_lbl = d_str
        day_type = day_types.get(d_str, "")
        bg, fg = THEME_TAG_COLORS.get(day_type, ("#F0F0F0", "#777777"))
        icon = THEME_ICONS.get(day_type, "")
        label = day_type.replace("_", " ").title() if day_type else ""
        header_html += (
            f'<th><span class="day-label">{d_lbl}</span>'
            f'<span class="theme-tag" style="background:{bg};color:{fg};">'
            f'{icon} {label}</span></th>')
    header_html += '</tr>'
    all_slots = set()
    for d_str in plan_dates:
        all_slots.update(plan.get(d_str, {}).keys())
    sorted_slots = sorted(all_slots, key=slot_sort_key)
    body_html = ''
    for slot_id in sorted_slots:
        body_html += f'<tr><td>{display_label_for_slot_id(slot_id)}</td>'
        for d_str in plan_dates:
            item = plan.get(d_str, {}).get(slot_id, "")
            is_nv = slot_id in nonveg.get(d_str, ())
            body_html += f'<td>{format_item_html(item, is_nonveg=is_nv)}</td>'
        body_html += '</tr>'
    return (
        f'<div class="menu-table-wrap"><table class="menu-table">'
        f'<thead>{header_html}</thead><tbody>{body_html}</tbody></table></div>'
    )


def _sanitize_sheet_title(name: str, used: set) -> str:
    """Excel sheet titles: <=31 chars, none of ``[]:*?/\\``, and unique."""
    title = re.sub(r'[\[\]:*?/\\]', ' ', str(name or "Counter")).strip()[:31] or "Counter"
    base, n = title, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _download_filename(blocks: list, client_name: str) -> str:
    """``menu_<client>_<date-range>.xlsx`` — dates span every non-empty block."""
    safe_client = re.sub(r'[^A-Za-z0-9]+', '_', client_name or "client").strip('_') or "client"
    dates = sorted({d for b in blocks if b.get("plan") for d in b.get("plan_dates", [])})
    if not dates:
        return f"menu_{safe_client}.xlsx"
    span = dates[0] if dates[0] == dates[-1] else f"{dates[0]}_to_{dates[-1]}"
    return f"menu_{safe_client}_{span}.xlsx"


def _plan_xlsx(blocks: list, client_name: str) -> bytes:
    """Formatted workbook — one sheet per counter, with bold bordered headers
    and non-veg dishes in red. Works for single (one sheet) and multi."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="131313")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="131313")
    header_fill = PatternFill("solid", fgColor="FEBF34")
    title_font = Font(bold=True, size=13, color="131313")
    nonveg_font = Font(color="C40D1B", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set = set()
    real = [b for b in blocks if b.get("plan")]

    for b in real:
        dates = b["plan_dates"]
        nonveg = b.get("nonveg") or {}
        ws = wb.create_sheet(_sanitize_sheet_title(b["name"], used_titles))

        # Title row (counter name).
        ws.cell(row=1, column=1, value=b["name"]).font = title_font

        # Header row.
        headers = ["Category"] + [_date_label(d) for d in dates]
        for col, text in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=text)
            c.font, c.fill, c.border = header_font, header_fill, border
            c.alignment = wrap

        # Body rows.
        slots = sorted(
            {s for d in dates for s in b["plan"].get(d, {})},
            key=slot_sort_key,
        )
        for r, slot_id in enumerate(slots, start=3):
            cat = ws.cell(row=r, column=1, value=display_label_for_slot_id(slot_id))
            cat.font, cat.border, cat.alignment = header_font, border, wrap
            for col, d in enumerate(dates, start=2):
                item = format_item_for_ui(b["plan"].get(d, {}).get(slot_id, ""))
                cell = ws.cell(row=r, column=col, value=item)
                cell.border, cell.alignment = border, wrap
                if slot_id in nonveg.get(d, ()):
                    cell.font = nonveg_font

        # Column widths: Category a bit wider, dates comfortable.
        ws.column_dimensions["A"].width = 18
        for col in range(2, len(dates) + 2):
            ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 22

    if not real:  # nothing generated yet — hand back an empty but valid file
        wb.create_sheet("Menu")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _pool_warnings_expander(block: dict) -> None:
    warns = block.get("pool_warnings") or []
    if warns:
        with st.expander(f"Pool warnings ({len(warns)})", expanded=False):
            for w in warns:
                st.markdown(
                    f'<div class="pool-warn-bar">&#9888; {html.escape(str(w))}</div>',
                    unsafe_allow_html=True)


def _render_regen_expander(api, block_index: int, counter_index: int,
                           key_ns: str) -> None:
    """Regenerate-cells panel for one plan block; mutates
    st.session_state.plan_blocks[block_index] in place and reruns."""
    b = st.session_state.plan_blocks[block_index]
    plan, plan_dates, day_types = b["plan"], b["plan_dates"], b["day_types"]
    with st.expander("Regenerate cells"):
        st.caption("Pick slots to replace with fresh items.")
        regen_selections = {}
        cols_per_row = min(len(plan_dates), 3) or 1
        cols = st.columns(cols_per_row)
        for i, d_str in enumerate(plan_dates):
            try:
                d_lbl = dt.date.fromisoformat(d_str).strftime("%a %d %b")
            except ValueError:
                d_lbl = d_str
            day_type = day_types.get(d_str, "")
            bg, fg = THEME_TAG_COLORS.get(day_type, ("#F0F0F0", "#777777"))
            icon = THEME_ICONS.get(day_type, "")
            label = day_type.replace("_", " ").title() if day_type else ""
            with cols[i % cols_per_row]:
                st.markdown(
                    f'<div class="regen-day-header">{d_lbl} '
                    f'<span class="theme-tag" style="background:{bg};color:{fg};'
                    f'font-size:0.6rem;">{icon} {label}</span></div>',
                    unsafe_allow_html=True)
                day_map = plan.get(d_str, {})

                def _fmt(slot_id, _dm=day_map):
                    cur = format_item_for_ui(_dm.get(slot_id, ""))
                    lbl = display_label_for_slot_id(slot_id)
                    return f"{lbl} — {cur}" if cur else lbl

                day_slots = sorted(day_map.keys(), key=slot_sort_key)
                selected = st.multiselect(
                    f"Slots for {d_str}", day_slots, format_func=_fmt,
                    key=f"regen_{key_ns}_{d_str}", label_visibility="collapsed")
                if selected:
                    regen_selections[d_str] = selected

        if st.button("Regenerate Selected", type="primary",
                     key=f"regen_btn_{key_ns}"):
            if not regen_selections:
                st.warning("Select at least one cell.")
                return
            old_snap = {
                (d, s): plan.get(d, {}).get(s, "")
                for d, slots in regen_selections.items() for s in slots
            }
            with st.spinner("Regenerating..."):
                try:
                    result = api.regenerate(
                        client_name=st.session_state.client_name,
                        base_plan=plan, replace_slots=regen_selections,
                        start_date=plan_dates[0], num_days=len(plan_dates),
                        time_limit_seconds=_PLANNING_TIME_LIMIT_SECONDS,
                        counter_index=counter_index)
                    solution = result.get("solution", {})
                    flat_regen, regen_day_types = flatten_api_solution(solution)
                    new_plan = flat_regen if flat_regen else plan
                    b["plan"] = new_plan
                    if regen_day_types:
                        b["day_types"] = regen_day_types
                    b["plan_dates"] = sorted(new_plan.keys())
                    if flat_regen:
                        b["nonveg"] = nonveg_slots_from_solution(solution)

                    diffs = []
                    for (d, s), old_raw in old_snap.items():
                        op = format_item_for_ui(old_raw)
                        np = format_item_for_ui(new_plan.get(d, {}).get(s, ""))
                        if op == np:
                            continue
                        try:
                            dl = dt.date.fromisoformat(d).strftime("%a %d %b")
                        except ValueError:
                            dl = d
                        diffs.append({
                            "kind": "regen", "counter": b["name"], "day": dl,
                            "slot": display_label_for_slot_id(s),
                            "old": op, "new": np,
                        })
                    if diffs:
                        st.session_state.changes_log.extend(diffs)
                        b["source"] = "modified"
                        st.session_state.plan_source = "modified"
                    st.rerun()
                except (ConnectionError, OSError, ValueError, RuntimeError) as e:
                    st.error(f"Regeneration failed: {e}")


def _render_changes_log() -> None:
    log = st.session_state.get("changes_log") or []
    if not log:
        return
    with st.expander("Changes log", expanded=True):
        for entry in log:
            if isinstance(entry, dict) and entry.get("kind") == "regen":
                ctr = entry.get("counter")
                ctr_html = (
                    f'<span class="log-slot">{html.escape(ctr)}</span>'
                    f'<span class="log-sep">&middot;</span>' if ctr else ''
                )
                st.markdown(
                    '<div class="log-entry log-diff">'
                    f'{ctr_html}'
                    f'<span class="log-day">{html.escape(entry["day"])}</span>'
                    f'<span class="log-sep">&middot;</span>'
                    f'<span class="log-slot">{html.escape(entry["slot"])}</span>'
                    f'<span class="log-sep">&middot;</span>'
                    f'<span class="log-old">{html.escape(entry["old"] or "(empty)")}</span>'
                    '<span class="log-arrow">&rarr;</span>'
                    f'<span class="log-new">{html.escape(entry["new"] or "(empty)")}</span>'
                    '</div>',
                    unsafe_allow_html=True)
            else:
                text = entry.get("text", "") if isinstance(entry, dict) else str(entry)
                st.markdown(
                    f'<div class="log-entry">{html.escape(text)}</div>',
                    unsafe_allow_html=True)


def _client_counter_names(api, name: str):
    """Return (mode, [counter names], city) for a client; degrade to single."""
    try:
        cfg = api.get_client_config(name)
        counters = cfg.get("counters") or []
        names = [(c.get("name") or f"Counter {i + 1}")
                 for i, c in enumerate(counters)] or ["Counter 1"]
        return cfg.get("counter_mode", "single"), names, cfg.get("city")
    except Exception:
        return "single", ["Counter 1"], None


# ---------------------------------------------------------------------------
# Generate
# ---------------------------------------------------------------------------
if generate_clicked:
    if not selected_client or selected_client == "(no clients)":
        st.warning("Select a valid client first.")
    else:
        mode, counter_names, city = _client_counter_names(client, selected_client)
        st.session_state.client_name = selected_client
        st.session_state.client_city = city
        st.session_state.plan_mode = mode
        st.session_state.changes_log = []
        st.session_state.plan_source = None
        st.session_state.rule_diagnostics = []
        st.session_state.diagnostics_summary = None

        if mode != "multi":
            # Single-cuisine: saved-plan replay, else solve (unchanged flow).
            try:
                saved = client.get_saved_plan(
                    client_name=selected_client,
                    start_date=start_date.isoformat(), num_days=num_days)
            except (ConnectionError, OSError, ValueError, RuntimeError) as e:
                st.warning(f"Couldn't check saved history ({e}); generating fresh.")
                saved = {"exists": False}

            if saved.get("exists"):
                with st.spinner(f"Loading saved plan for {selected_client}..."):
                    blk = _flatten_result(saved)
                    blk["name"] = counter_names[0]
                    blk["source"] = "history"
                    st.session_state.plan_blocks = [blk]
                    st.session_state.plan_source = "history"
                    st.rerun()
            else:
                with st.spinner(f"Generating plan for {selected_client}..."):
                    try:
                        result = client.plan(
                            client_name=selected_client,
                            start_date=start_date.isoformat(), num_days=num_days,
                            time_limit_seconds=_PLANNING_TIME_LIMIT_SECONDS)
                        blk = _flatten_result(result)
                        blk["name"] = counter_names[0]
                        st.session_state.plan_blocks = [blk]
                        st.session_state.plan_source = "solver"
                        st.session_state.rule_diagnostics = result.get("rule_diagnostics") or []
                        st.session_state.diagnostics_summary = result.get("summary")
                        st.rerun()
                    except RuleDiagnosticsBlockedError as e:
                        st.session_state.plan_blocks = [{
                            "name": counter_names[0], "plan": {}, "plan_dates": [],
                            "day_types": {}, "pool_warnings": [],
                            "source": "preflight_blocked",
                            "error": str(e) or "Pre-flight blocked",
                        }]
                        st.session_state.plan_source = "preflight_blocked"
                        st.session_state.rule_diagnostics = e.diagnostics or []
                        st.session_state.diagnostics_summary = e.summary or None
                        st.rerun()
                    except (ConnectionError, OSError, ValueError, RuntimeError) as e:
                        st.error(f"Generation failed: {e}")
        else:
            # Multi-cuisine: solve each counter independently. Divide the
            # time budget across counters so total wall-clock stays bounded.
            per_limit = max(45, _PLANNING_TIME_LIMIT_SECONDS // max(1, len(counter_names)))
            blocks = []
            with st.spinner(
                f"Generating {len(counter_names)} counters for {selected_client}..."
            ):
                for i, cname in enumerate(counter_names):
                    try:
                        result = client.plan(
                            client_name=selected_client,
                            start_date=start_date.isoformat(), num_days=num_days,
                            time_limit_seconds=per_limit, counter_index=i)
                        blk = _flatten_result(result)
                        blk["name"] = cname
                    except RuleDiagnosticsBlockedError as e:
                        blk = {"name": cname, "plan": {}, "plan_dates": [],
                               "day_types": {}, "pool_warnings": [],
                               "source": "preflight_blocked",
                               "error": str(e) or "Pre-flight blocked for this counter"}
                    except (ConnectionError, OSError, ValueError, RuntimeError) as e:
                        blk = {"name": cname, "plan": {}, "plan_dates": [],
                               "day_types": {}, "pool_warnings": [],
                               "source": "error", "error": str(e)}
                    blocks.append(blk)
            st.session_state.plan_blocks = blocks
            st.session_state.plan_source = "solver"
            st.rerun()

# ---------------------------------------------------------------------------
# Display
# ---------------------------------------------------------------------------
_blocks = st.session_state.get("plan_blocks") or []
_plan_mode = st.session_state.get("plan_mode", "single")

_render_diagnostics_expander(
    st.session_state.get("rule_diagnostics") or [],
    st.session_state.get("diagnostics_summary"),
)

# Single-counter pre-flight block: no table, just the CTA + diagnostics above.
if (
    _plan_mode != "multi" and _blocks
    and _blocks[0].get("source") == "preflight_blocked"
    and not _blocks[0].get("plan")
):
    st.warning(
        "Pre-flight diagnostics found a guaranteed failure for these dates. "
        "Fix the issues above (or change the dates / client) and try again.")
    st.stop()

if _blocks and any(b.get("plan") for b in _blocks):
    dates_union = sorted({d for b in _blocks for d in b.get("plan_dates", [])})
    total_items = sum(
        1 for b in _blocks for d in b.get("plan_dates", [])
        for s in b["plan"].get(d, {}) if b["plan"][d].get(s)
    )
    # Metric cards (Counters is always shown now).
    cards = [
        ("Client", html.escape(st.session_state.client_name or "")),
    ]
    _city = st.session_state.get("client_city")
    if _city:
        cards.append(("City", html.escape(_city)))
    cards += [
        ("Counters", str(len(_blocks))),
        ("Days", str(len(dates_union))),
    ]
    if _plan_mode != "multi":
        b0 = _blocks[0]
        slots_per_day = len({s for d in b0["plan_dates"] for s in b0["plan"].get(d, {})})
        cards.append(("Slots per day", str(slots_per_day)))
    cards.append(("Total items", str(total_items)))
    st.markdown(
        '<div class="metrics-grid">' + ''.join(
            f'<div class="metric-card"><div class="metric-label">{lbl}</div>'
            f'<div class="metric-value">{val}</div></div>'
            for lbl, val in cards
        ) + '</div>', unsafe_allow_html=True)

    if _plan_mode == "multi":
        # Shared Save / Download / Clear-all bar.
        sc1, sc2, sc3, _sc = st.columns([1.3, 1, 1, 3])
        with sc1:
            if st.button("Save All to History", type="primary",
                         key="multi_save_btn", use_container_width=True):
                payload = [{"name": b["name"], "week_plan": b["plan"]}
                           for b in _blocks if b.get("plan")]
                try:
                    client.save(client_name=st.session_state.client_name,
                                week_start=dates_union[0], counters=payload)
                    for b in _blocks:
                        if b.get("plan"):
                            b["source"] = "history"
                    st.session_state.plan_source = "history"
                    st.toast("All counters saved to history", icon="✅")
                except (ConnectionError, OSError, ValueError, RuntimeError) as e:
                    st.error(f"Save failed: {e}")
        with sc2:
            st.download_button(
                "Download Excel",
                data=_plan_xlsx(_blocks, st.session_state.client_name),
                file_name=_download_filename(_blocks, st.session_state.client_name),
                mime=_XLSX_MIME, key="multi_dl_btn", use_container_width=True)
        with sc3:
            if st.button("Clear All", key="multi_clear_btn",
                         use_container_width=True):
                st.session_state.plan_blocks = []
                st.session_state.changes_log = []
                st.session_state.plan_source = None
                st.rerun()

        tabs = st.tabs([b["name"] for b in _blocks])
        for i, (tab, b) in enumerate(zip(tabs, _blocks)):
            with tab:
                if b.get("error") and not b.get("plan"):
                    st.warning(f"&#9888; {b['name']}: {b['error']}")
                    continue
                _pool_warnings_expander(b)
                st.markdown(
                    _menu_table_html(b["plan"], b["plan_dates"], b["day_types"],
                                     b.get("nonveg")),
                    unsafe_allow_html=True)
                st.markdown("")
                cc1, cc2, _cc = st.columns([1, 1, 4])
                with cc2:
                    if st.button("Clear", key=f"clear_c{i}",
                                 use_container_width=True):
                        b["plan"], b["plan_dates"], b["day_types"] = {}, [], {}
                        b["nonveg"] = {}
                        st.rerun()
                _render_regen_expander(client, i, i, f"c{i}")
    else:
        b = _blocks[0]
        _pool_warnings_expander(b)
        st.markdown(
            _menu_table_html(b["plan"], b["plan_dates"], b["day_types"],
                             b.get("nonveg")),
            unsafe_allow_html=True)
        st.markdown("")
        c1, c2, c3, _c = st.columns([1, 1, 1, 3])
        with c1:
            if st.button("Save to History", key="planner_save_btn",
                         use_container_width=True):
                try:
                    client.save(client_name=st.session_state.client_name,
                                week_start=b["plan_dates"][0], week_plan=b["plan"])
                    st.session_state.plan_source = "history"
                    b["source"] = "history"
                    st.toast("Plan saved to history", icon="✅")
                except (ConnectionError, OSError, ValueError, RuntimeError) as e:
                    st.error(f"Save failed: {e}")
        with c2:
            st.download_button(
                "Download Excel",
                data=_plan_xlsx(_blocks, st.session_state.client_name),
                file_name=_download_filename(_blocks, st.session_state.client_name),
                mime=_XLSX_MIME, key="planner_download_xlsx_btn",
                use_container_width=True)
        with c3:
            if st.button("Clear", key="planner_clear_btn",
                         use_container_width=True):
                st.session_state.plan_blocks = []
                st.session_state.changes_log = []
                st.session_state.plan_source = None
                st.session_state.rule_diagnostics = []
                st.session_state.diagnostics_summary = None
                st.rerun()
        _render_regen_expander(client, 0, 0, "single")

    _render_changes_log()

else:
    st.markdown("""<div class="empty-state">
        <div class="empty-icon">&#127835;</div>
        <h3>No menu plan yet</h3>
        <p>Select a client and click <b>Generate Menu Plan</b><br>in the sidebar to get started.</p>
    </div>""", unsafe_allow_html=True)
