"""Pure render / export helpers for the planner.

Extracted from ``app.py`` so the entry-point script stays a thin Streamlit
flow and this logic (menu-table HTML, the Excel workbook, filenames, the
API-response → block shape) is importable and unit-testable on its own. None
of these functions touch Streamlit — they take plain data and return
strings / bytes / dicts.
"""

from __future__ import annotations

import datetime as dt
import io
import re
from typing import Dict, List, Optional

from ui.formatters import (
    THEME_TAG_COLORS,
    THEME_ICONS,
    display_label_for_slot_id,
    flatten_api_solution,
    format_item_for_ui,
    format_item_html,
    nonveg_slots_from_solution,
    slot_sort_key,
)

# MIME type for the .xlsx download.
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def flatten_result(result: dict) -> dict:
    """Turn a /plan or /saved-plan response into a plan "block"."""
    solution = result.get("solution", {})
    flat, day_types = flatten_api_solution(solution)
    return {
        "plan": flat,
        "plan_dates": sorted(flat.keys()),
        "day_types": day_types,
        "nonveg": nonveg_slots_from_solution(solution),
        "pool_warnings": result.get("pool_warnings", []),
        "source": "solver",
        "error": None,
    }


def date_label(d_str: str) -> str:
    try:
        return dt.date.fromisoformat(d_str).strftime("%a %d %b")
    except ValueError:
        return d_str


def menu_table_html(plan: dict, plan_dates: list, day_types: dict,
                    nonveg: Optional[dict] = None) -> str:
    nonveg = nonveg or {}
    header_html = '<tr><th>Category</th>'
    for d_str in plan_dates:
        d_lbl = date_label(d_str)
        day_type = day_types.get(d_str, "")
        bg, fg = THEME_TAG_COLORS.get(day_type, ("#F0F0F0", "#777777"))
        icon = THEME_ICONS.get(day_type, "")
        label = day_type.replace("_", " ").title() if day_type else ""
        header_html += (
            f'<th><span class="day-label">{d_lbl}</span>'
            f'<span class="theme-tag" style="background:{bg};color:{fg};">'
            f'{icon} {label}</span></th>')
    header_html += '</tr>'
    all_slots = set()
    for d_str in plan_dates:
        all_slots.update(plan.get(d_str, {}).keys())
    sorted_slots = sorted(all_slots, key=slot_sort_key)
    body_html = ''
    for slot_id in sorted_slots:
        body_html += f'<tr><td>{display_label_for_slot_id(slot_id)}</td>'
        for d_str in plan_dates:
            item = plan.get(d_str, {}).get(slot_id, "")
            is_nv = slot_id in nonveg.get(d_str, ())
            body_html += f'<td>{format_item_html(item, is_nonveg=is_nv)}</td>'
        body_html += '</tr>'
    return (
        f'<div class="menu-table-wrap"><table class="menu-table">'
        f'<thead>{header_html}</thead><tbody>{body_html}</tbody></table></div>'
    )


def sanitize_sheet_title(name: str, used: set) -> str:
    """Excel sheet titles: <=31 chars, none of ``[]:*?/\\``, and unique."""
    title = re.sub(r'[\[\]:*?/\\]', ' ', str(name or "Counter")).strip()[:31] or "Counter"
    base, n = title, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def download_filename(blocks: list, client_name: str) -> str:
    """``menu_<client>_<date-range>.xlsx`` — dates span every non-empty block."""
    safe_client = re.sub(r'[^A-Za-z0-9]+', '_', client_name or "client").strip('_') or "client"
    dates = sorted({d for b in blocks if b.get("plan") for d in b.get("plan_dates", [])})
    if not dates:
        return f"menu_{safe_client}.xlsx"
    span = dates[0] if dates[0] == dates[-1] else f"{dates[0]}_to_{dates[-1]}"
    return f"menu_{safe_client}_{span}.xlsx"


def plan_xlsx(blocks: List[Dict], client_name: str) -> bytes:
    """Formatted workbook — one sheet per counter, with bold bordered headers
    and non-veg dishes in red. Works for single (one sheet) and multi."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="131313")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="131313")
    header_fill = PatternFill("solid", fgColor="FEBF34")
    title_font = Font(bold=True, size=13, color="131313")
    nonveg_font = Font(color="C40D1B", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set = set()
    real = [b for b in blocks if b.get("plan")]

    for b in real:
        dates = b["plan_dates"]
        nonveg = b.get("nonveg") or {}
        ws = wb.create_sheet(sanitize_sheet_title(b["name"], used_titles))

        # Title row (counter name).
        ws.cell(row=1, column=1, value=b["name"]).font = title_font

        # Header row.
        headers = ["Category"] + [date_label(d) for d in dates]
        for col, text in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=text)
            c.font, c.fill, c.border = header_font, header_fill, border
            c.alignment = wrap

        # Body rows.
        slots = sorted(
            {s for d in dates for s in b["plan"].get(d, {})},
            key=slot_sort_key,
        )
        for r, slot_id in enumerate(slots, start=3):
            cat = ws.cell(row=r, column=1, value=display_label_for_slot_id(slot_id))
            cat.font, cat.border, cat.alignment = header_font, border, wrap
            for col, d in enumerate(dates, start=2):
                item = format_item_for_ui(b["plan"].get(d, {}).get(slot_id, ""))
                cell = ws.cell(row=r, column=col, value=item)
                cell.border, cell.alignment = border, wrap
                if slot_id in nonveg.get(d, ()):
                    cell.font = nonveg_font

        # Column widths: Category a bit wider, dates comfortable.
        ws.column_dimensions["A"].width = 18
        for col in range(2, len(dates) + 2):
            ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 22

    if not real:  # nothing generated yet — hand back an empty but valid file
        wb.create_sheet("Menu")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
