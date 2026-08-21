#!/usr/bin/env python3
"""Import Corning Chakan (Pune)'s nine-week master menu into the Pune ontology.

Source: `data/raw/source_workbooks/corning_chakan_pune_menu.xlsx` — nine weekly
sheets, one column per day, with the category in column A. The row layout is
**identical on every sheet**, and this is the first client menu imported into a
city other than Bangalore or Chennai, so it is also the first Maharashtrian
list: amti, patodi, thecha, matki, bharali wangi, fodanich waran.

**Lunch and Dinner only.** Breakfast (poha, misal pav, sabudana wada), Evening
Snacks and Midnight Snacks are separate services the tool does not plan — the
same call the Booking, Citrix and Stripe imports made. Importing them would put
a wada pav in the day's starter slot.

Four things specific to this workbook:

* **The salad block is a salad BAR.** Its first row carries a composed dish some
  days (`Papadi Chaat`, `Dahi Wada`) and a bare vegetable on others (`Tomato`,
  `Iceberg Lettuce`); the five rows beneath it are always components — beetroot,
  zucchini, boiled peanuts, cottage cheese. Components are ingredients a diner
  assembles, not dishes a slot can serve, and a menu printing "Tomato" as the
  day's salad is the `remove_generic_rows.py` problem arriving by a new route.
  So continuation rows are dropped wholesale (the `+` keys below are absent from
  `CATEGORY_MAP`, which is what skips them) and the labelled row is filtered
  against `SALAD_BAR_COMPONENTS`. The chaats it does carry are re-filed to
  `starter`, where this ontology already files every chaat, dhokla and samosa —
  which is also the pool Pune most needed: it had 7 starters.

* **Three printed rows each serve two slots**, so each is re-filed by name — a
  dish called a soup belongs in `soup` however its row is labelled. `DAL` holds
  a `Rassam` and a `Cream of Burnt Garlic Soup`; `DESSERT/SOUP` is half sweets
  and half soups; `SOUP / TETRA PACK` mixes soups with `Amul Sweet Lassi`,
  `Butter Milk` and `Masala Milk`, which are welcome drinks.

* **The chutney row names the ingredient, not the dish.** It writes `Tomato` and
  `Tomato Chutney`, `Jawas` and `Jawas Chutney`, on different days for the same
  thing. Under a row labelled CHUTNEY a cell saying "Tomato" means tomato
  chutney, so the category word is appended when it is missing — otherwise the
  fold sees two dishes and the ontology gains a row called `tomato`.

* **One unlabelled block.** The `10th to 16th Aug` sheet carries an Independence
  Day menu in the Saturday column below the grid (`Tiranga Pulao`, `Tiranga
  Burfi`), with no row labels at all. Position cannot settle a slot there, so
  the course comes from the dish NAME and anything the name cannot settle is
  skipped and reported. A blank row is what separates it from the grid above,
  so that is what this treats as the boundary: a label does not carry across
  one.

Pool token: every row is `common`, which is the Pune convention — the file is
the whole Pune universe (`docs/pune_rulebook.md`), and Pune is not in
`FULL_POOL_CITIES`, so a dish tagged `Corning Chakan` would be invisible to
every Pune client including this one.

Idempotent: re-running adds nothing and re-tags nothing.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))  # sibling scripts
ROOT = Path(__file__).resolve().parent.parent
SOURCE = (ROOT / "data" / "raw" / "source_workbooks"
          / "corning_chakan_pune_menu.xlsx")
PUNE = ROOT / "data" / "raw" / "city_items" / "pune.xlsx"

from menu_import import (  # noqa: E402  (needs the sys.path line above)
    ImportSpec,
    course_from_name,
    is_placeholder,
    norm,
    refile_lentils,
    run_import,
    to_item,
)

#: Pune's file is the whole city universe, so every row is `common`.
CLIENT_TOKEN = "common"

#: Section headings. A row whose label is one of these and whose day cells are
#: all empty opens a block; the labels beneath it belong to that block.
BLOCK_HEADERS = {"lunch", "dinner", "evening snacks", "midnight snacks"}

#: Rows that are not a category at all. `build` checks this BEFORE the category
#: map, so a label wanted for even one block must stay out of it — the salad
#: bar's continuation rows are dropped by their `+` key being absent from
#: `CATEGORY_MAP`, not by listing `salad` here.
SKIP_LABELS = {"date", "day", "breakfast", "dressings", "tea", "accompaniment",
               "veg na", "evening snacks", "midnight snacks"}

#: Courses the unlabelled Independence Day block can land in. Its key is
#: rewritten to `SPECIAL||<course>` at parse time from the dish NAME, so `build`
#: still gets a real course for every dish it is handed.
SPECIAL_COURSES = ("veg_gravy", "veg_dry", "dal", "rice", "bread", "papad",
                   "salad", "chutney", "dessert", "starter", "curd_side",
                   "soup", "healthy_rice")

#: '<block>||<label>' -> the app's course_type. A key absent here is skipped,
#: which is how the salad-bar continuation rows (`SALAD+`) and every non-plated
#: block are dropped.
CATEGORY_MAP = {
    "LUNCH||dry veg": "veg_dry",
    "LUNCH||gravy veg": "veg_gravy",
    "LUNCH||dal": "dal",
    "LUNCH||rice": "rice",
    "LUNCH||chapati": "bread",
    "LUNCH||papad": "papad",
    "LUNCH||salad": "salad",
    "LUNCH||chutney": "chutney",
    "LUNCH||dessert/soup": "dessert",
    "LUNCH||soup / tetra pack": "soup",
    "DINNER||dry veg": "veg_dry",
    "DINNER||gravy veg": "veg_gravy",
    "DINNER||dal": "dal",
    "DINNER||rice": "rice",
    "DINNER||chapati": "bread",
    "DINNER||papad": "papad",
    "DINNER||salad": "salad",
    "DINNER||chutney": "chutney",
    "DINNER||dessert/soup": "dessert",
    "MIDNIGHT SNACKS||soup / tetra pack": "soup",
}
#: The unlabelled Independence Day block, keyed by the course its dish name
#: resolves to (see `parse`).
CATEGORY_MAP.update({f"SPECIAL||{c}": c for c in SPECIAL_COURSES})

#: The salad bar's components: produce and protein toppings, not dishes. Taken
#: from what this workbook actually writes in those rows.
SALAD_BAR_COMPONENTS = {
    "american_corn", "beet_root", "beetroot", "bell_peppers", "boiled_babycorn",
    "boiled_brocoli", "boiled_broccoli", "boiled_chana", "boiled_chickpeas",
    "boiled_fresh_beans", "boiled_moong", "boiled_peanuts", "capsicum",
    "caroot", "carrot", "cauliflower", "cottage_cheese", "cucumber",
    "french_beans", "purpal_cabbage", "purple_cabbage", "raddish", "radish",
    "roasted_sweet_potatoes", "soya_chunks", "tofu", "zucchini",
    "iceberg_lettuce", "tomato", "onion", "lettuce",
}

#: Category-named rows this must never create — `remove_generic_rows.py` deletes
#: `salad` and `sweet` from Pune, and a const-slot word alone is not a dish.
SKIP_ITEMS = {
    "salad", "sweet", "chutney", "papad", "pickle", "dressings", "tea",
    "plain_rice", "steamed_rice", "steamed_plain_rice", "white_rice",
    "same_as_lunch", "shared_separately", "cut_fruits",
}

#: Placeholders this source writes that the shared list does not carry.
_PLACEHOLDER_TEXT = {
    "same as lunch", "same as dinner", "same as lunch and dinner",
    "same as lunch & dinner", "shared separately", "same as lunch/dinner",
}

#: Words that make a dish a soup whatever row it is printed under.
_SOUP_WORDS = {"soup", "shorba", "broth", "chowder", "rassam", "rasam"}

#: Welcome drinks printed under `SOUP / TETRA PACK`.
_DRINK_WORDS = {"lassi", "milk", "buttermilk", "chaas", "chaach", "sherbet",
                "juice", "smoothie"}

#: Chaats and fried snacks the salad row carries; this ontology files every one
#: of them as a starter.
_STARTER_WORDS = {"chaat", "chat", "wada", "vada", "dhokla", "samosa",
                  "kachori", "bhel", "pakoda", "pakora", "tikki"}

#: Salad dressings, which the unlabelled festival block carries with the bar
#: components. Scoped to that block — `sauce` alone is far too broad to filter
#: on generally, `hot_garlic_sauce` being a real dish elsewhere.
_DRESSING_WORDS = {"dressing", "vinaigrette", "mayo", "sauce"}

#: The four unlabelled festival dishes `course_from_name` cannot place. It knows
#: dal, rice, bread, starter and dessert words, but not the `papad` and
#: `chutney` const slots, nor the gravy FORMS (`pasanda`, `dum aloo`) — a
#: Banarasi dum aloo and a paneer pasanda are both gravies. Adjudicated one at
#: a time rather than by widening the shared helper, which six other importers
#: read.
FESTIVAL_COURSES = {
    "paneer_gulnaz_pasanda": "veg_gravy",
    "stuffed_dum_aloo_banarasi": "veg_gravy",
    "colorful_papad": "papad",
    "dahi_wada": "starter",
    "tomato_chutney": "chutney",
}


#: Whole-name fixes this source needs. A token rule cannot express them:
#: `ACHARI PANEER R` ends in a stray column marker, `R PAPAD` abbreviates
#: "roasted", and the last four are one dish written twice in the same workbook
#: with the words in the other order or a form word added — which the fold
#: cannot settle, both spellings arriving new in the same import.
SOURCE_ALIASES = {
    "achari_paneer_r": "achari_paneer",
    "r_papad": "roasted_papad",
    "kashmiri_dum_aloo": "dum_aloo_kashmiri",
    "palak_lasooni": "lasooni_palak",
    "malai_kofta_curry": "malai_kofta",
    "babycorn_mushroom_masala": "babycorn_mushroom",
    "cream_of_burnt_garlic_soup": "burnt_garlic_soup",
    # `VEG` adds nothing in an all-veg city, and the dish is the same one.
    "veg_sweet_corn_soup": "sweet_corn_soup",
    # Pune already carries this dish under the spelling its own list uses.
    # Without the alias the import adds a second motichoor laddu.
    "motichoor_laddoo": "moti_chur_laddu",
    "motichoor_laddu": "moti_chur_laddu",
    # "Mysore Pak in Pure Ghee" IS the `ghee_mysore_pak` the ontology already
    # carries. Under the source's wording the dish is unique, so no other city
    # can settle its `dessert_form` and the token vote has too few mysore paks
    # to reach threshold — it arrives as a dessert the variety rule cannot group.
    "mysore_pak_in_pure_ghee": "ghee_mysore_pak",
    # A gravy that arrives by splitting "CHAPATI + MIX VEG"; without this the
    # one-dish-one-category pass claims it for `bread` and the ontology gains a
    # mixed-veg gravy servable as the day's roti.
    "mix_veg": "mix_veg",
}

#: Dishes that must land in a given course whatever row they were printed under.
ITEM_COURSE = {"mix_veg": "veg_gravy", "pav_bhaji": "veg_dry"}


def clean_name(raw: str) -> str:
    item = to_item(raw, drop_parentheticals=True)
    return SOURCE_ALIASES.get(item, item)


def _is_placeholder(text: str) -> bool:
    return (is_placeholder(text)
            or norm(text).strip().lower() in _PLACEHOLDER_TEXT)


def parse_source(verbose: bool = True) -> dict:
    """{'<block>||<label>': [raw dish, …]} for the plated lunch/dinner rows."""
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    out: dict = defaultdict(list)
    for ws in wb.worksheets:
        block = "LUNCH"
        label = None
        seen_label_rows = 0
        for row in range(1, ws.max_row + 1):
            printed = norm(ws.cell(row, 1).value).strip()
            cells = [norm(ws.cell(row, col).value).strip()
                     for col in range(2, 9)]
            if not printed and not any(cells):
                # A blank row ends whatever block was open. It is the only
                # thing separating the Independence Day menu from the grid.
                label = None
                continue
            if printed:
                low = printed.lower()
                if low in BLOCK_HEADERS and not any(cells):
                    block = printed.upper()
                    label = None
                    continue
                label = low
                seen_label_rows += 1
                key = f"{block}||{label}"
            elif label is not None:
                key = f"{block}||{label}+"          # a continuation row
            else:
                key = "SPECIAL||festival"           # no label anywhere above
            if label is not None and label in SKIP_LABELS \
                    and key not in CATEGORY_MAP:
                continue
            for cell in cells:
                if not _is_placeholder(cell):
                    out[key].append(cell)
    if verbose:
        for key in sorted(out):
            mark = "" if key in CATEGORY_MAP else "   (skipped)"
            print(f"  {key:<38}{len(out[key]):>4}{mark}")
    return dict(out)


def _clean_chutney(text: str) -> str:
    """`Tomato` under a CHUTNEY row means tomato chutney."""
    item = to_item(text)
    return item if item.endswith("chutney") else f"{item}_chutney"


def parse(verbose: bool = False) -> dict:
    """The spec's parse hook: the grid, with two source-specific readings.

    The chutney rows get the category word they omit, and the unlabelled
    festival block is re-keyed to whatever course each dish NAME resolves to.
    """
    raw = parse_source(verbose=verbose)

    for key, dishes in list(raw.items()):
        if not key.endswith("||chutney"):
            continue
        named = []
        for d in dishes:
            item = to_item(d)
            if item == "pickle" or "raita" in item or "dahi" in item:
                named.append(d)            # re-filed by name below
            else:
                named.append(_clean_chutney(d).replace("_", " "))
        raw[key] = named

    festival = raw.pop("SPECIAL||festival", [])
    unresolved = []
    for d in festival:
        item = to_item(d)
        if item in SALAD_BAR_COMPONENTS or set(item.split("_")) & _DRESSING_WORDS:
            continue                       # the bar and its dressings again
        course = FESTIVAL_COURSES.get(item) or course_from_name(item)
        if course in SPECIAL_COURSES:
            raw.setdefault(f"SPECIAL||{course}", []).append(d)
        else:
            unresolved.append(d)
    if unresolved:
        print(f"  ! {len(unresolved)} unlabelled festival dish(es) skipped — "
              f"the name does not say which slot: {sorted(unresolved)}")
    return raw


def refile(item: str, course: str) -> str:
    """Three printed rows serve two slots each; the name decides which."""
    words = set(item.split("_"))

    if item in ITEM_COURSE:
        return ITEM_COURSE[item]

    if words & _SOUP_WORDS:
        return "rasam" if words & {"rassam", "rasam"} else "soup"

    if course == "soup" and words & _DRINK_WORDS:
        return "welcome_drink"

    if course == "chutney":
        if "pickle" in words:
            return "pickle"
        if words & {"raita", "dahi"}:
            return "curd_side"

    if course == "salad" and words & _STARTER_WORDS:
        return "starter"

    if course == "veg_dry" and words & {"noodles", "noodle"}:
        return "rice"

    if course == "veg_gravy" and "dry" in words:
        return "veg_dry"

    return refile_lentils(item, course)


SPEC = ImportSpec(
    client_token=CLIENT_TOKEN,
    city_path=PUNE,
    parse=parse,
    category_map=CATEGORY_MAP,
    skip_labels=SKIP_LABELS,
    refile=refile,
    clean_name=clean_name,
    skip_items=SKIP_ITEMS | SALAD_BAR_COMPONENTS,
    split_combos=True,
)


def main(dry_run: bool = False):
    return run_import(SPEC, dry_run=dry_run)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--show-grid", action="store_true",
                    help="print what each printed row contributed")
    args = ap.parse_args()
    if args.show_grid:
        parse_source(verbose=True)
    else:
        main(dry_run=args.dry_run)
