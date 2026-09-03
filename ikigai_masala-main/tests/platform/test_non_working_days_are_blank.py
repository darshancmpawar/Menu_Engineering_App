"""A day the client does not work is a BLANK column, not a missing one.

The horizon spans `num_days` weekdays (or calendar days for a weekend site).
A client with a restricted `working_days` list serves only some of them —
Clario is Mon-Thu, Piramel Mon/Tue/Thu, Quince Wed/Thu/Fri — and the other days
used to be filtered out before the solve and never came back. A 5-day horizon
from Monday returned FOUR days for Clario: the table had no Friday at all, the
gap closed up, and "5 days" quietly meant a different number of columns per
client.

Now the solver still plans only the served days, but the menu is rendered over
the whole span, so Friday is there and empty. Three claims:

* the span is what gets rendered, and the blank days are the unserved ones,
* the solve is unchanged — no rule sees an extra day, and no dish is planned
  for one,
* and an empty column is LABELLED, because otherwise it is indistinguishable
  from a day the solver failed on.
"""

from __future__ import annotations

import datetime as dt

import pytest

from tests.fake_supabase import FakeSupabase

MONDAY = '2026-09-07'          # a Monday
#: name -> (working_days, the weekday names that must come back blank)
RESTRICTED = {
    'Clario': (['monday', 'tuesday', 'wednesday', 'thursday'], {'Fri'}),
    'Quince': (['wednesday', 'thursday', 'friday'], {'Mon', 'Tue'}),
}


@pytest.fixture(scope='module')
def api():
    import src.db as db_mod
    import api.app as api_app
    from tests.client_fixtures import CLIENTS

    wanted = set(RESTRICTED) | {'Amadeus Pune'}
    rows = [dict(c) for c in CLIENTS if c['name'] in wanted]
    db_mod._sb_client = FakeSupabase(seed={
        'clients': rows, 'app_settings': [],
        'menu_history': [], 'week_signatures': [],
    })
    api_app._client_loader = None
    api_app.reset_caches()
    api_app.app.config['TESTING'] = True
    return api_app


def _plan(api, client, days=5, start=MONDAY):
    from api.rate_limit import reset_for_tests
    reset_for_tests()
    r = api.app.test_client().post('/api/v1/plan', json={
        'client_name': client, 'start_date': start,
        'num_days': days, 'time_limit': 40,
    })
    assert r.status_code == 200, r.get_json()
    return (r.get_json() or {}).get('solution') or {}


def _weekday(date_key):
    return dt.date.fromisoformat(date_key).strftime('%a')


class TestTheHorizonKeepsItsShape:
    @pytest.mark.parametrize('client', sorted(RESTRICTED))
    def test_the_span_is_rendered_not_the_served_subset(self, api, client):
        """5 asked for, 5 columns back — whatever the client's week is."""
        sol = _plan(api, client)
        assert len(sol) == 5, sorted(sol)

    @pytest.mark.parametrize('client', sorted(RESTRICTED))
    def test_the_blank_days_are_exactly_the_unserved_ones(self, api, client):
        _working, expect_blank = RESTRICTED[client]
        sol = _plan(api, client)
        blank = {_weekday(d) for d, v in sol.items() if not (v.get('items') or {})}
        assert blank == expect_blank, sorted(sol)

    @pytest.mark.parametrize('client', sorted(RESTRICTED))
    def test_every_served_day_actually_has_a_menu(self, api, client):
        """The other half — widening the span must not have emptied a real
        day."""
        _working, expect_blank = RESTRICTED[client]
        sol = _plan(api, client)
        for d, v in sol.items():
            if _weekday(d) in expect_blank:
                continue
            assert (v.get('items') or {}), f'{d} came back empty'

    def test_an_unrestricted_client_is_unchanged(self, api):
        """All but three clients have no `working_days` list, and for them
        nothing about this may differ."""
        sol = _plan(api, 'Amadeus Pune')
        assert len(sol) == 5
        assert all(v.get('items') for v in sol.values())
        assert all(v.get('is_working_day') for v in sol.values())

    def test_ten_days_is_two_working_weeks(self, api):
        """A horizon of 10 spans two Mon-Fri weeks, so Clario gets two blank
        Fridays and eight menus — not ten menus and not eight columns."""
        sol = _plan(api, 'Clario', days=10)
        assert len(sol) == 10
        blank = [d for d, v in sol.items() if not (v.get('items') or {})]
        assert {_weekday(d) for d in blank} == {'Fri'}
        assert len(blank) == 2, sorted(blank)


class TestTheBlankDayIsMarked:
    @pytest.mark.parametrize('client', sorted(RESTRICTED))
    def test_it_carries_is_working_day_false(self, api, client):
        _working, expect_blank = RESTRICTED[client]
        sol = _plan(api, client)
        for d, v in sol.items():
            assert v.get('is_working_day') is (_weekday(d) not in expect_blank), d

    def test_it_carries_no_theme(self, api):
        """A cuisine tag over an empty column says the kitchen was supposed to
        cook something and didn't."""
        sol = _plan(api, 'Clario')
        off = [v for d, v in sol.items() if _weekday(d) == 'Fri']
        assert off and not off[0]['day_type'] and not off[0]['theme']

    def test_a_served_day_still_carries_its_theme(self, api):
        sol = _plan(api, 'Clario')
        on = [v for d, v in sol.items() if _weekday(d) == 'Mon']
        assert on and on[0]['day_type']


class TestTheUiRendersItAsAColumn:
    def test_the_table_has_a_labelled_off_column(self, api):
        from ui.planner_view import flatten_result, menu_table_html
        block = flatten_result({'solution': _plan(api, 'Clario')})
        assert len(block['plan_dates']) == 5
        assert block['off_days'] == {'2026-09-11'}
        html = menu_table_html(block['plan'], block['plan_dates'],
                               block['day_types'], block.get('nonveg'),
                               block.get('off_days'))
        assert 'Not served' in html
        assert 'day-off' in html

    def test_an_unrestricted_client_gets_no_off_column(self, api):
        from ui.planner_view import flatten_result, menu_table_html
        block = flatten_result({'solution': _plan(api, 'Amadeus Pune')})
        assert block['off_days'] == set()
        html = menu_table_html(block['plan'], block['plan_dates'],
                               block['day_types'], block.get('nonveg'),
                               block.get('off_days'))
        assert 'Not served' not in html

    def test_the_excel_export_keeps_the_blank_column(self, api):
        """A printed menu with the day silently removed is the same bug in a
        different file format."""
        import io
        from openpyxl import load_workbook
        from ui.planner_view import flatten_result, plan_xlsx
        block = flatten_result({'solution': _plan(api, 'Clario')})
        wb = load_workbook(io.BytesIO(
            plan_xlsx([{**block, 'name': 'Main'}], 'Clario')))
        ws = wb['Main']
        headers = [c.value for c in ws[2]]
        assert len(headers) == 6, headers          # Category + 5 days
        assert any('11' in str(h) for h in headers), headers

    def test_off_days_defaults_to_empty_for_an_old_payload(self):
        """`is_working_day` is new; a cached or replayed body without it must
        not turn every day into a blank one."""
        from ui.formatters import off_days_from_solution
        assert off_days_from_solution(
            {'2026-09-07': {'theme': 'Mix', 'items': {}}}) == set()
        assert off_days_from_solution({}) == set()
        assert off_days_from_solution(None) == set()


class TestTheSolveIsUnchanged:
    def test_no_dish_is_planned_for_an_unserved_day(self, api):
        sol = _plan(api, 'Quince')
        for d, v in sol.items():
            if _weekday(d) in RESTRICTED['Quince'][1]:
                assert not (v.get('items') or {}), d

    def test_a_weekly_rule_still_counts_calendar_weeks(self, api):
        """The blank day must not become a day a rule can count. Clario over 10
        days spans two ISO weeks, and that is what a `max_per_week` rule sees —
        the two blank Fridays add nothing."""
        sol = _plan(api, 'Clario', days=10)
        served = [dt.date.fromisoformat(d) for d, v in sol.items()
                  if (v.get('items') or {})]
        weeks = {(d.isocalendar()[0], d.isocalendar()[1]) for d in served}
        assert len(weeks) == 2, sorted(weeks)
